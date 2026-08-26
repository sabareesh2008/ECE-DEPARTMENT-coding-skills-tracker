from __future__ import annotations

import csv
import os
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, time as dt_time
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# SETTINGS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

STUDENTS_FILE = BASE_DIR / "students.csv"
LIVE_CSV = BASE_DIR / "LiveData.csv"
HISTORY_CSV = BASE_DIR / "History.csv"
DAILY_ACTIVITY_CSV = BASE_DIR / "DailyActivity.csv"
STUDENTS_XLSX = BASE_DIR / "Students.xlsx"

SUPABASE_URL = (
    os.getenv("SUPABASE_URL", "")
    .strip()
    .replace("%0A", "")
    .replace("%0a", "")
    .rstrip("/")
)
SUPABASE_SERVICE_ROLE_KEY = os.getenv(
    "SUPABASE_SERVICE_ROLE_KEY",
    "",
).strip()

LEETCODE_URL = "https://leetcode.com/graphql"

RECENT_SUBMISSION_LIMIT = 2000
IST = ZoneInfo("Asia/Kolkata")

# Check up to 10 LeetCode profiles at the same time.
MAX_WORKERS = 8

ALLOWED_SECTIONS = (
    "ECE A",
    "ECE B",
    "ECE C",
    "ECE D",
    "ECE E",
    "ECE F",
)


LEETCODE_QUERY = """
query getUserProfile($username: String!, $limit: Int!) {
  matchedUser(username: $username) {
    username
    submitStatsGlobal {
      acSubmissionNum {
        difficulty
        count
        submissions
      }
      totalSubmissionNum {
        difficulty
        count
        submissions
      }
    }
    userCalendar {
      submissionCalendar
    }
  }

  recentAcSubmissionList(
    username: $username,
    limit: $limit
  ) {
    title
    titleSlug
    timestamp
  }
}
"""


# ============================================================
# HELPERS
# ============================================================

def clean(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def empty_profile(status: str) -> dict[str, Any]:
    return {
        "total_solved": 0,
        "easy": 0,
        "medium": 0,
        "hard": 0,
        "submissions": 0,
        "solved_today": 0,
        "last_7_days": 0,
        "last_14_days": 0,
        "last_30_days": 0,
        "last_7_days_submissions": 0,
        "last_problem": "",
        "last_solved": "",
        "status": status,
        "recent_submissions": [],
    }


def get_stat(
    statistics: list[dict[str, Any]],
    difficulty: str,
    field: str = "count",
) -> int:
    for statistic in statistics:
        if statistic.get("difficulty") == difficulty:
            return int(statistic.get(field, 0) or 0)

    return 0


def _submission_datetime(
    submission: dict[str, Any],
) -> datetime | None:
    timestamp = submission.get("timestamp")

    if timestamp in (None, ""):
        return None

    try:
        return datetime.fromtimestamp(
            int(timestamp),
            tz=IST,
        )
    except (TypeError, ValueError, OSError):
        return None


def _submission_key(
    submission: dict[str, Any],
) -> str:
    return (
        clean(submission.get("titleSlug"))
        or clean(submission.get("title"))
    )


def _calendar_window_start(days: int) -> datetime:
    """
    Calendar-day window in IST.

    7 Days = today + previous 6 calendar dates.
    14 Days = today + previous 13 calendar dates.
    30 Days = today + previous 29 calendar dates.
    """
    return datetime.combine(
        ist_today() - timedelta(days=days - 1),
        dt_time.min,
        tzinfo=IST,
    )


def _normalize_recent_accepted(
    recent_submissions: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Keep only valid accepted-submission records and sort newest first.
    """
    cleaned_rows: list[dict[str, Any]] = []
    now = ist_now() + timedelta(minutes=10)

    for submission in recent_submissions or []:
        submitted_at = _submission_datetime(submission)
        key = _submission_key(submission)

        if (
            submitted_at is None
            or not key
            or submitted_at > now
        ):
            continue

        row = dict(submission)
        row["_submitted_at"] = submitted_at
        row["_problem_key"] = key
        cleaned_rows.append(row)

    cleaned_rows.sort(
        key=lambda item: item["_submitted_at"],
        reverse=True,
    )

    return cleaned_rows


def unique_accepted_problems_in_window(
    recent_submissions: list[dict[str, Any]],
    start_time: datetime,
) -> int:
    """
    Count DISTINCT problems with at least one accepted submission
    inside the requested window.

    Re-submitting the same problem five times still counts as one problem
    for that window.
    """
    solved: set[str] = set()

    for submission in recent_submissions:
        submitted_at = submission.get("_submitted_at")
        problem_key = submission.get("_problem_key")

        if (
            isinstance(submitted_at, datetime)
            and problem_key
            and submitted_at >= start_time
        ):
            solved.add(str(problem_key))

    return len(solved)


def accepted_feed_covers_window(
    recent_submissions: list[dict[str, Any]],
    accepted_submission_total: int,
    start_time: datetime,
) -> bool:
    """
    Decide whether the returned accepted-submission feed fully covers
    a requested time window.

    Coverage is proven when either:
      1) every lifetime accepted submission is present in the returned list, or
      2) the oldest returned accepted submission is at/before the window start.

    This prevents silently treating a truncated recent list as exact.
    """
    accepted_submission_total = max(
        0,
        safe_int(accepted_submission_total),
    )

    if accepted_submission_total == 0:
        return True

    if not recent_submissions:
        return False

    if len(recent_submissions) >= accepted_submission_total:
        return True

    oldest = recent_submissions[-1].get("_submitted_at")

    return bool(
        isinstance(oldest, datetime)
        and oldest <= start_time
    )


def submission_count_calendar_days(
    submission_calendar: Any,
    days: int,
) -> int:
    """
    Count ALL submission attempts over calendar days in IST.
    """
    if not submission_calendar:
        return 0

    try:
        calendar = (
            json.loads(submission_calendar)
            if isinstance(submission_calendar, str)
            else submission_calendar
        )
    except (TypeError, ValueError, json.JSONDecodeError):
        return 0

    if not isinstance(calendar, dict):
        return 0

    start_date = (
        ist_today() - timedelta(days=days - 1)
    )

    end_date = ist_today()

    total = 0

    for timestamp, count in calendar.items():
        try:
            submitted_date = datetime.fromtimestamp(
                int(timestamp),
                tz=IST,
            ).date()

            if start_date <= submitted_date <= end_date:
                total += max(0, int(count or 0))

        except (TypeError, ValueError, OSError):
            continue

    return total


def validate_window_order(
    today_count: int,
    seven_count: int,
    fourteen_count: int,
    thirty_count: int,
) -> bool:
    """
    Because all four metrics come from nested accepted-submission windows,
    this invariant must always hold.
    """
    return (
        0
        <= today_count
        <= seven_count
        <= fourteen_count
        <= thirty_count
    )


# ============================================================
# LEETCODE FETCH
# ============================================================

def fetch_leetcode(username: str) -> dict[str, Any]:
    """
    Fetch one public LeetCode profile.

    IMPORTANT DEFINITION USED BY CODEMETRIX:
    "Solved in N days" = number of DISTINCT problem titles that have at least
    one ACCEPTED submission in that calendar-day window.

    We do NOT mix this with cumulative-total snapshot deltas.
    Mixing those two definitions was the reason 7d could be non-zero while
    14d became zero.
    """
    if not username:
        return empty_profile("Username missing")

    request_body = {
        "operationName": "getUserProfile",
        "query": LEETCODE_QUERY,
        "variables": {
            "username": username,
            "limit": RECENT_SUBMISSION_LIMIT,
        },
    }

    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Origin": "https://leetcode.com",
        "Referer": f"https://leetcode.com/u/{username}/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/151.0.0.0 Safari/537.36"
        ),
    }

    last_error = ""

    # Retry transient rate-limit/server/network failures.
    for attempt in range(1, 4):
        try:
            response = requests.post(
                LEETCODE_URL,
                json=request_body,
                headers=headers,
                timeout=35,
            )

            if response.status_code == 429:
                last_error = "HTTP 429 rate limited"

                if attempt < 3:
                    import time
                    time.sleep(attempt * 2)
                    continue

                return empty_profile(last_error)

            if response.status_code >= 500:
                last_error = f"HTTP {response.status_code}"

                if attempt < 3:
                    import time
                    time.sleep(attempt * 2)
                    continue

                return empty_profile(last_error)

            if response.status_code != 200:
                return empty_profile(
                    f"HTTP {response.status_code}"
                )

            try:
                response_data = response.json()
            except ValueError:
                return empty_profile(
                    "Invalid JSON response"
                )

            graphql_errors = response_data.get("errors") or []

            calendar_permission_error = False
            fatal_messages = []

            for error in graphql_errors:
                message = str(
                    error.get(
                        "message",
                        "GraphQL error",
                    )
                )

                lowered = message.lower()

                if (
                    "calendar" in lowered
                    and (
                        "permission" in lowered
                        or "no permission" in lowered
                    )
                ):
                    calendar_permission_error = True
                else:
                    fatal_messages.append(message)

            data = response_data.get("data", {}) or {}

            # GraphQL may return useful matchedUser/recentAcSubmissionList data
            # together with a field-level calendar permission error.
            if fatal_messages and not data:
                return empty_profile(
                    " | ".join(fatal_messages)
                )
            matched_user = data.get("matchedUser")

            if matched_user is None:
                return empty_profile("User not found")

            submit_stats = (
                matched_user
                .get("submitStatsGlobal", {})
                or {}
            )

            accepted_stats = (
                submit_stats
                .get("acSubmissionNum", [])
                or []
            )

            total_stats = (
                submit_stats
                .get("totalSubmissionNum", [])
                or []
            )

            raw_recent = (
                data.get("recentAcSubmissionList", [])
                or []
            )

            recent_submissions = (
                _normalize_recent_accepted(
                    raw_recent
                )
            )

            user_calendar = (
                matched_user.get("userCalendar")
                if isinstance(matched_user, dict)
                else None
            )

            calendar_available = (
                not calendar_permission_error
                and isinstance(user_calendar, dict)
            )

            submission_calendar = (
                user_calendar.get(
                    "submissionCalendar",
                    "{}",
                )
                if calendar_available
                else "{}"
            )

            easy = get_stat(
                accepted_stats,
                "Easy",
            )

            medium = get_stat(
                accepted_stats,
                "Medium",
            )

            hard = get_stat(
                accepted_stats,
                "Hard",
            )

            total_solved_api = get_stat(
                accepted_stats,
                "All",
            )

            difficulty_sum = (
                easy + medium + hard
            )

            # Difficulty totals are an independent consistency check.
            # Prefer All normally, but never allow a lower impossible total.
            total_solved = max(
                total_solved_api,
                difficulty_sum,
            )

            accepted_submission_total = get_stat(
                accepted_stats,
                "All",
                "submissions",
            )

            total_submissions = get_stat(
                total_stats,
                "All",
                "submissions",
            )

            # Some LeetCode responses use count rather than submissions for
            # totalSubmissionNum. Use the larger non-negative value.
            total_submissions = max(
                total_submissions,
                get_stat(
                    total_stats,
                    "All",
                    "count",
                ),
            )

            now = ist_now()

            today_start = datetime.combine(
                ist_today(),
                dt_time.min,
                tzinfo=IST,
            )

            seven_start = (
                _calendar_window_start(7)
            )

            fourteen_start = (
                _calendar_window_start(14)
            )

            thirty_start = (
                _calendar_window_start(30)
            )

            solved_today = (
                unique_accepted_problems_in_window(
                    recent_submissions,
                    today_start,
                )
            )

            last_7_days = (
                unique_accepted_problems_in_window(
                    recent_submissions,
                    seven_start,
                )
            )

            last_14_days = (
                unique_accepted_problems_in_window(
                    recent_submissions,
                    fourteen_start,
                )
            )

            last_30_days = (
                unique_accepted_problems_in_window(
                    recent_submissions,
                    thirty_start,
                )
            )

            coverage = {
                "today":
                    accepted_feed_covers_window(
                        recent_submissions,
                        accepted_submission_total,
                        today_start,
                    ),
                "7d":
                    accepted_feed_covers_window(
                        recent_submissions,
                        accepted_submission_total,
                        seven_start,
                    ),
                "14d":
                    accepted_feed_covers_window(
                        recent_submissions,
                        accepted_submission_total,
                        fourteen_start,
                    ),
                "30d":
                    accepted_feed_covers_window(
                        recent_submissions,
                        accepted_submission_total,
                        thirty_start,
                    ),
            }

            if not validate_window_order(
                solved_today,
                last_7_days,
                last_14_days,
                last_30_days,
            ):
                return empty_profile(
                    "Calculation validation failed"
                )

            last_problem = ""
            last_solved = ""

            if recent_submissions:
                latest = recent_submissions[0]

                last_problem = clean(
                    latest.get("title")
                )

                submitted_at = latest.get(
                    "_submitted_at"
                )

                if isinstance(
                    submitted_at,
                    datetime,
                ):
                    last_solved = (
                        submitted_at.strftime(
                            "%Y-%m-%d %H:%M:%S IST"
                        )
                    )

            result = {
                "total_solved": total_solved,
                "easy": easy,
                "medium": medium,
                "hard": hard,
                "submissions": total_submissions,
                "solved_today": solved_today,
                "last_7_days": last_7_days,
                "last_14_days": last_14_days,
                "last_30_days": last_30_days,
                "last_7_days_submissions":
                    submission_count_calendar_days(
                        submission_calendar,
                        7,
                    ),
                "last_problem": last_problem,
                "last_solved": last_solved,
                "status": "Success",
                "recent_submissions":
                    recent_submissions,
                "window_coverage": coverage,
                "accepted_submission_total":
                    accepted_submission_total,
                "recent_accepted_returned":
                    len(recent_submissions),
                "calendar_available":
                    calendar_available,
                "calendar_note":
                    (
                        ""
                        if calendar_available
                        else "Calendar unavailable"
                    ),
            }

            return result

        except requests.Timeout:
            last_error = "Request timeout"

        except requests.ConnectionError as error:
            last_error = (
                f"Connection error: {error}"
            )

        except requests.RequestException as error:
            last_error = (
                f"Network error: {error}"
            )

        except Exception as error:
            last_error = (
                f"Unexpected error: {error}"
            )

        if attempt < 3:
            import time
            time.sleep(attempt * 2)

    return empty_profile(
        last_error or "Unknown fetch error"
    )


# ============================================================
# SAFE CSV WRITE
# ============================================================

def atomic_csv_write(
    dataframe: pd.DataFrame,
    destination: Path,
) -> None:
    temporary_file = destination.with_suffix(
        ".temporary.csv"
    )

    dataframe.to_csv(
        temporary_file,
        index=False,
        encoding="utf-8-sig",
    )

    os.replace(
        temporary_file,
        destination,
    )


# ============================================================
# INPUT / SUPABASE
# ============================================================

def sync_students_from_supabase() -> None:
    """Download the authoritative student directory from Supabase."""

    if (
        not SUPABASE_URL
        or not SUPABASE_SERVICE_ROLE_KEY
    ):
        print(
            "Supabase secrets not set; "
            "using local students.csv"
        )
        return

    url = f"{SUPABASE_URL}/rest/v1/students"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization":
            f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Accept": "application/json",
    }

    params = {
        "select": (
            "register_number,"
            "student_name,"
            "leetcode_username,"
            "section,"
            "created_at"
        ),
        "order": "created_at.asc",
    }

    response = requests.get(
        url,
        headers=headers,
        params=params,
        timeout=30,
    )

    response.raise_for_status()

    rows = response.json()

    frame = pd.DataFrame(
        [
            {
                "Register Number":
                    clean(
                        row.get(
                            "register_number"
                        )
                    ),
                "Student Name":
                    clean(
                        row.get(
                            "student_name"
                        )
                    ),
                "LeetCode Username":
                    clean(
                        row.get(
                            "leetcode_username"
                        )
                    ),
                "Section":
                    clean(
                        row.get("section")
                    )
                    or "ECE E",
            }
            for row in rows
        ]
    )

    if frame.empty:
        frame = pd.DataFrame(
            columns=[
                "Register Number",
                "Student Name",
                "LeetCode Username",
                "Section",
            ]
        )

    atomic_csv_write(
        frame,
        STUDENTS_FILE,
    )

    print(
        f"Synced {len(frame)} student(s) "
        "from Supabase"
    )


def write_students_excel(
    students: pd.DataFrame,
) -> None:
    """Create an Excel copy automatically."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = [
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "Section",
        "LeetCode Link",
    ]

    ws.append(headers)

    for _, student in students.iterrows():
        username = clean(
            student["LeetCode Username"]
        )

        ws.append(
            [
                clean(
                    student[
                        "Register Number"
                    ]
                ),
                clean(
                    student[
                        "Student Name"
                    ]
                ),
                username,
                clean(
                    student["Section"]
                ),
                (
                    "https://leetcode.com/u/"
                    f"{username}/"
                ),
            ]
        )

    header_fill = PatternFill(
        "solid",
        fgColor="2563EB",
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center"
        )

    widths = {
        "A": 20,
        "B": 28,
        "C": 28,
        "D": 14,
        "E": 48,
    }

    for column, width in widths.items():
        ws.column_dimensions[
            column
        ].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(
        STUDENTS_XLSX
    )


def read_students() -> pd.DataFrame:
    sync_students_from_supabase()

    if not STUDENTS_FILE.exists():
        raise FileNotFoundError(
            f"students.csv not found: "
            f"{STUDENTS_FILE}"
        )

    students = pd.read_csv(
        STUDENTS_FILE,
        dtype=str,
        keep_default_na=False,
    )

    # Backward compatibility:
    # old local CSV files did not have Section.
    if "Section" not in students.columns:
        students["Section"] = "ECE E"

    required_columns = {
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "Section",
    }

    missing = (
        required_columns
        .difference(students.columns)
    )

    if missing:
        raise ValueError(
            "Missing students.csv columns: "
            + ", ".join(
                sorted(missing)
            )
        )

    students = (
        students
        .dropna(how="all")
        .copy()
    )

    for column in required_columns:
        students[column] = (
            students[column]
            .apply(clean)
        )

    students.loc[
        students["Section"] == "",
        "Section",
    ] = "ECE E"

    invalid_sections = sorted(
        set(students["Section"])
        .difference(
            ALLOWED_SECTIONS
        )
    )

    if invalid_sections:
        raise ValueError(
            "Invalid section value(s): "
            + ", ".join(
                invalid_sections
            )
        )

    students = students[
        (
            students[
                "Student Name"
            ]
            != ""
        )
        & (
            students[
                "LeetCode Username"
            ]
            != ""
        )
    ].copy()

    write_students_excel(
        students
    )

    return students


# ============================================================
# COMPLETED-DAY ROLLING 7 / 30 DAY COUNTS
# ============================================================

def safe_int(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def ist_now() -> datetime:
    """Single source of truth for CodeMetrix tracker time."""
    return datetime.now(IST)


def ist_today() -> date:
    """Calendar date in Asia/Kolkata, not the GitHub runner's UTC date."""
    return ist_now().date()


def load_daily_activity_file() -> pd.DataFrame:
    columns = [
        "Date",
        "Section",
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "Problems Solved",
        "Solved That Day",
        "Source",
        "Exact",
    ]

    if not DAILY_ACTIVITY_CSV.exists():
        return pd.DataFrame(columns=columns)

    try:
        frame = pd.read_csv(
            DAILY_ACTIVITY_CSV,
            dtype=str,
            keep_default_na=False,
        )
    except pd.errors.EmptyDataError:
        return pd.DataFrame(columns=columns)

    for column in columns:
        if column not in frame.columns:
            frame[column] = ""

    return frame[columns].copy()


def latest_previous_snapshot(
    history: pd.DataFrame,
    register_number: str,
) -> dict[str, Any] | None:
    if history.empty:
        return None

    student_history = history[
        history["Register Number"].astype(str) == str(register_number)
    ].copy()

    if student_history.empty:
        return None

    student_history["_date"] = pd.to_datetime(
        student_history["Date"],
        errors="coerce",
    )

    today = pd.Timestamp(ist_today())

    student_history = student_history[
        student_history["_date"] < today
    ].dropna(subset=["_date"])

    if student_history.empty:
        return None

    row = (
        student_history
        .sort_values("_date")
        .iloc[-1]
    )

    return row.to_dict()


def solved_on_date(
    activity: pd.DataFrame,
    register_number: str,
    target_date: date,
) -> int:
    if activity.empty:
        return 0

    matches = activity[
        (activity["Register Number"].astype(str) == str(register_number))
        & (activity["Date"].astype(str) == target_date.isoformat())
    ]

    if matches.empty:
        return 0

    return safe_int(matches.iloc[-1]["Solved That Day"])



def _student_daily_total_snapshots(
    history: pd.DataFrame,
    register_number: str,
    username: str = "",
    current_total: int | None = None,
) -> pd.DataFrame:
    """
    Return trustworthy cumulative Problems Solved snapshots.

    Protection rules:
    - same register number
    - same LeetCode username when username history exists
    - Status must be Success when Status exists
    - valid date + non-negative total
    - historical total cannot exceed current total
    - obvious E/M/H-vs-total corruption is rejected
    - decreasing cumulative rows are discarded
    - leading zero bootstrap rows are ignored once a later positive snapshot
      exists; this prevents a failed first fetch of 0 from becoming a fake
      152-problem 7-day increase

    Old rolling 7/14/30 columns are never used as calculation inputs.
    """
    columns = ["_date", "_total", "_order"]

    if history is None or history.empty:
        return pd.DataFrame(columns=columns)

    required = {
        "Date",
        "Register Number",
        "Problems Solved",
    }

    if not required.issubset(history.columns):
        return pd.DataFrame(columns=columns)

    frame = history[
        history["Register Number"].astype(str).str.strip()
        == str(register_number).strip()
    ].copy()

    if frame.empty:
        return pd.DataFrame(columns=columns)

    # A register number can survive while the linked LeetCode username changes.
    # Never mix the old account's cumulative totals into the new account.
    if username and "LeetCode Username" in frame.columns:
        wanted = str(username).strip().lower()

        username_series = (
            frame["LeetCode Username"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )

        username_rows = frame[username_series == wanted].copy()

        if not username_rows.empty:
            frame = username_rows
        else:
            return pd.DataFrame(columns=columns)

    # A failed request must never become a cumulative baseline.
    if "Status" in frame.columns:
        status = (
            frame["Status"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )

        frame = frame[status == "success"].copy()

    if frame.empty:
        return pd.DataFrame(columns=columns)

    frame["_date"] = pd.to_datetime(
        frame["Date"],
        errors="coerce",
    ).dt.date

    frame["_total"] = pd.to_numeric(
        frame["Problems Solved"],
        errors="coerce",
    )

    frame["_order"] = range(len(frame))

    frame = frame.dropna(
        subset=["_date", "_total"]
    )

    frame = frame[
        frame["_total"] >= 0
    ].copy()

    if current_total is not None:
        current_total = max(0, safe_int(current_total))

        # A historical cumulative total above the current cumulative total is
        # impossible for the same LeetCode account.
        frame = frame[
            frame["_total"] <= current_total
        ].copy()

    if frame.empty:
        return pd.DataFrame(columns=columns)

    # Independent corruption check using the difficulty totals.
    difficulty_columns = {"Easy", "Medium", "Hard"}

    if difficulty_columns.issubset(frame.columns):
        easy = pd.to_numeric(frame["Easy"], errors="coerce")
        medium = pd.to_numeric(frame["Medium"], errors="coerce")
        hard = pd.to_numeric(frame["Hard"], errors="coerce")

        difficulty_sum = easy + medium + hard
        known_difficulty = (
            easy.notna()
            & medium.notna()
            & hard.notna()
        )

        valid_breakdown = (
            ~known_difficulty
            | (difficulty_sum == frame["_total"])
        )

        frame = frame[valid_breakdown].copy()

    if frame.empty:
        return pd.DataFrame(columns=columns)

    frame = (
        frame
        .sort_values(["_date", "_order"])
        .drop_duplicates(
            subset=["_date"],
            keep="last",
        )
        .sort_values("_date")
        .reset_index(drop=True)
    )

    # If a profile later has a positive cumulative total, old leading zero
    # snapshots are too risky to use as an exact baseline. They may have come
    # from an early API/profile fetch failure that was historically saved as
    # Success. Dropping them can only reduce certainty; it cannot fabricate a
    # huge rolling value.
    positive_positions = frame.index[
        frame["_total"] > 0
    ].tolist()

    if positive_positions:
        first_positive = positive_positions[0]
        frame = frame.loc[first_positive:].copy()

    if frame.empty:
        return pd.DataFrame(columns=columns)

    # Cumulative solved count cannot decrease for the same account.
    # Discard the decreasing row instead of carrying a previous value forward,
    # because carrying it forward would falsely create an "exact" snapshot.
    trusted_rows = []
    last_total = None

    for _, row in frame.iterrows():
        total = safe_int(row["_total"])

        if last_total is None or total >= last_total:
            trusted_rows.append(row)
            last_total = total

    if not trusted_rows:
        return pd.DataFrame(columns=columns)

    trusted = pd.DataFrame(trusted_rows)

    return trusted[columns].reset_index(drop=True)


def _history_window_value(
    history: pd.DataFrame,
    register_number: str,
    username: str,
    current_total: int,
    days: int,
) -> tuple[int | None, str]:
    """
    Calculate a rolling NEW-problem count from trustworthy cumulative totals.

    HISTORY_EXACT:
      trustworthy snapshot exists exactly N calendar days ago.

    LOWER_BOUND:
      the exact boundary is absent, but a later trustworthy snapshot exists.
      This value is guaranteed, but may undercount the full window.

    MISSING:
      insufficient trustworthy history.
    """
    current_total = max(0, safe_int(current_total))

    frame = _student_daily_total_snapshots(
        history,
        register_number,
        username=username,
        current_total=current_total,
    )

    if frame.empty:
        return None, "MISSING"

    target_date = ist_today() - timedelta(days=days)

    exact = frame[
        frame["_date"] == target_date
    ]

    if not exact.empty:
        baseline = safe_int(
            exact.iloc[-1]["_total"]
        )

        value = current_total - baseline

        if value < 0 or value > current_total:
            return None, "MISSING"

        return value, "HISTORY_EXACT"

    after = frame[
        (frame["_date"] > target_date)
        & (frame["_date"] < ist_today())
    ]

    if not after.empty:
        baseline = safe_int(
            after.iloc[0]["_total"]
        )

        value = current_total - baseline

        if value < 0 or value > current_total:
            return None, "MISSING"

        return value, "LOWER_BOUND"

    return None, "MISSING"


def _choose_rolling_value(
    history_value: int | None,
    history_source: str,
    recent_value: int,
    recent_full: bool,
) -> tuple[int, str]:
    """
    Choose one trustworthy source.

    Important cross-check:
    If the recent accepted feed is proven complete for the window, the number
    of NEW problems from cumulative history cannot be greater than the number
    of distinct accepted problems seen in that complete window.

    Therefore, when:
      history exact = 152
      recent complete = 0
    the historical boundary is rejected as contaminated instead of publishing
    an impossible 152.
    """
    recent_value = max(0, safe_int(recent_value))

    if (
        history_source == "HISTORY_EXACT"
        and history_value is not None
    ):
        history_value = max(
            0,
            safe_int(history_value),
        )

        if recent_full and history_value > recent_value:
            return recent_value, "RECENT_FULL_HISTORY_REJECTED"

        return history_value, "HISTORY_EXACT"

    if recent_full:
        return recent_value, "RECENT_FULL"

    if (
        history_source == "LOWER_BOUND"
        and history_value is not None
    ):
        return (
            max(0, safe_int(history_value)),
            "LOWER_BOUND",
        )

    return 0, "INSUFFICIENT_HISTORY"


def calculate_rolling_metrics(
    previous_history: pd.DataFrame,
    register_number: str,
    username: str,
    profile: dict[str, Any],
) -> dict[str, Any]:
    """
    Final rolling NEW-problem calculation.

    Invariants enforced before output:
      0 <= Today <= 7D <= 14D <= 30D <= Total Solved

    A missing/partial source is never allowed to overwrite a stronger shorter
    window with an impossible smaller value.
    """
    current_total = max(
        0,
        safe_int(profile.get("total_solved")),
    )

    coverage = profile.get(
        "window_coverage",
        {},
    ) or {}

    values = {}

    for label, days, recent_key, coverage_key in [
        ("today", 1, "solved_today", "today"),
        ("7d", 7, "last_7_days", "7d"),
        ("14d", 14, "last_14_days", "14d"),
        ("30d", 30, "last_30_days", "30d"),
    ]:
        history_value, history_source = _history_window_value(
            previous_history,
            register_number,
            username,
            current_total,
            days,
        )

        value, source = _choose_rolling_value(
            history_value,
            history_source,
            profile.get(recent_key, 0),
            bool(coverage.get(coverage_key)),
        )

        values[label] = min(
            current_total,
            max(0, safe_int(value)),
        )

        values[f"{label}_source"] = source

    # Nested calendar windows.
    if values["7d"] < values["today"]:
        values["7d"] = values["today"]
        values["7d_source"] = "DERIVED_LOWER_BOUND"

    if values["14d"] < values["7d"]:
        values["14d"] = values["7d"]
        values["14d_source"] = "DERIVED_LOWER_BOUND"

    if values["30d"] < values["14d"]:
        values["30d"] = values["14d"]
        values["30d_source"] = "DERIVED_LOWER_BOUND"

    # Final hard guard. This should always pass after normalization.
    valid = (
        0
        <= values["today"]
        <= values["7d"]
        <= values["14d"]
        <= values["30d"]
        <= current_total
    )

    if not valid:
        raise RuntimeError(
            "Rolling metric integrity failure: "
            f"today={values['today']} "
            f"7d={values['7d']} "
            f"14d={values['14d']} "
            f"30d={values['30d']} "
            f"total={current_total}"
        )

    return values


def calculate_completed_day_counts(
    previous_history: pd.DataFrame,
    previous_activity: pd.DataFrame,
    register_number: str,
    current_total: int,
    solved_today: int,
    leetcode_7_days: int,
    leetcode_14_days: int,
    leetcode_30_days: int,
) -> tuple[int, int, int, str, int]:
    """
    Backward-compatible helper.

    Main tracker execution uses calculate_rolling_metrics() with the real
    username. This wrapper intentionally uses an empty username only for older
    internal callers/tests.
    """
    del previous_activity

    pseudo_profile = {
        "total_solved": current_total,
        "solved_today": solved_today,
        "last_7_days": leetcode_7_days,
        "last_14_days": leetcode_14_days,
        "last_30_days": leetcode_30_days,
        "window_coverage": {
            "today": False,
            "7d": False,
            "14d": False,
            "30d": False,
        },
    }

    metrics = calculate_rolling_metrics(
        previous_history,
        register_number,
        "",
        pseudo_profile,
    )

    return (
        metrics["7d"],
        metrics["14d"],
        metrics["30d"],
        ist_today().isoformat(),
        metrics["today"],
    )


def update_completed_daily_activity(
    previous_activity: pd.DataFrame,
    current_rows: list[dict[str, Any]],
) -> pd.DataFrame:
    columns = [
        "Date",
        "Section",
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "Problems Solved",
        "Solved That Day",
        "Source",
        "Exact",
    ]

    activity = previous_activity.copy()

    for column in columns:
        if column not in activity.columns:
            activity[column] = ""

    new_rows = []

    for row in current_rows:
        completed_date = row.get("_Completed Date", "")
        completed_solved = safe_int(
            row.get("_Completed Solved", 0)
        )
        source = str(
            row.get(
                "_Completed Source",
                "",
            )
        )

        if not completed_date:
            continue

        exact = source in {
            "HISTORY_EXACT",
            "RECENT_FULL",
        }

        if not activity.empty:
            activity = activity[
                ~(
                    (activity["Date"].astype(str) == str(completed_date))
                    & (
                        activity["Register Number"].astype(str)
                        == str(row["Register Number"])
                    )
                )
            ].copy()

        new_rows.append({
            "Date": completed_date,
            "Section": row["Section"],
            "Register Number": row["Register Number"],
            "Student Name": row["Student Name"],
            "LeetCode Username": row["LeetCode Username"],
            "Problems Solved": row["Problems Solved"],
            "Solved That Day": completed_solved,
            "Source": source,
            "Exact": "true" if exact else "false",
        })

    if new_rows:
        activity = pd.concat(
            [
                activity,
                pd.DataFrame(new_rows, columns=columns),
            ],
            ignore_index=True,
        )

    if not activity.empty:
        activity = activity.sort_values(
            by=["Date", "Section", "Register Number"],
            ascending=[True, True, True],
        ).reset_index(drop=True)

    return activity[columns]



# ============================================================
# DAILY CHALLENGE
# ============================================================

def supabase_headers() -> dict[str, str]:
    return {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }


def load_recent_challenges(days: int = 35) -> list[dict[str, Any]]:
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        return []

    start_date = (
        datetime.now(IST).date()
        - timedelta(days=days)
    ).isoformat()

    try:
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/daily_challenges",
            headers=supabase_headers(),
            params={
                "select": (
                    "id,challenge_date,problem_title,"
                    "problem_slug,problem_url,difficulty"
                ),
                "challenge_date": f"gte.{start_date}",
                "order": "challenge_date.asc",
            },
            timeout=30,
        )

        if not response.ok:
            print(
                "Daily challenge load warning: "
                f"{response.status_code} {response.text[:300]}"
            )
            return []

        return response.json()

    except requests.RequestException as error:
        print(
            "Daily challenge load network warning: "
            f"{error}"
        )
        return []


def challenge_completion(
    recent_submissions: list[dict[str, Any]],
    challenge: dict[str, Any],
) -> tuple[bool, str | None]:
    """
    Current rule:
    accepted submission must be on the challenge date in IST.
    """

    try:
        challenge_day = date.fromisoformat(
            str(challenge["challenge_date"])
        )
    except (KeyError, ValueError, TypeError):
        return False, None

    target_slug = clean(
        challenge.get("problem_slug")
    ).lower()

    if not target_slug:
        return False, None

    for submission in recent_submissions:
        if (
            clean(
                submission.get("titleSlug")
            ).lower()
            != target_slug
        ):
            continue

        timestamp = submission.get(
            "timestamp"
        )

        if not timestamp:
            continue

        try:
            solved_at = datetime.fromtimestamp(
                int(timestamp),
                tz=IST,
            )
        except (TypeError, ValueError, OSError):
            continue

        if solved_at.date() == challenge_day:
            return (
                True,
                solved_at.isoformat(),
            )

    return False, None


def save_challenge_result(
    challenge_id: int,
    register_number: str,
    completed: bool,
    completed_at: str | None,
) -> bool:
    """
    Robust challenge-result save:
    - GET then PATCH/POST (no on_conflict dependency)
    - Completed never becomes Pending
    - DB/network failures do not crash the LeetCode tracker
    """

    if (
        not SUPABASE_URL
        or not SUPABASE_SERVICE_ROLE_KEY
    ):
        return False

    register_number = clean(
        register_number
    )

    if not register_number:
        return False

    base_url = (
        f"{SUPABASE_URL}/rest/v1/"
        "daily_challenge_results"
    )

    headers = supabase_headers()

    try:
        lookup = requests.get(
            base_url,
            headers=headers,
            params={
                "select":
                    "id,completed,completed_at",
                "challenge_id":
                    f"eq.{challenge_id}",
                "register_number":
                    f"eq.{register_number}",
                "limit": "1",
            },
            timeout=30,
        )

        if not lookup.ok:
            print(
                "  Challenge lookup warning "
                f"[{register_number}]: "
                f"{lookup.status_code} "
                f"{lookup.text[:180]}"
            )
            return False

        existing_rows = lookup.json()

        existing_completed = False
        existing_completed_at = None

        if existing_rows:
            existing_completed = bool(
                existing_rows[0].get(
                    "completed"
                )
            )
            existing_completed_at = (
                existing_rows[0].get(
                    "completed_at"
                )
            )

        final_completed = (
            existing_completed
            or completed
        )

        final_completed_at = (
            existing_completed_at
            if existing_completed
            else completed_at
        )

        payload = {
            "completed":
                final_completed,
            "completed_at":
                final_completed_at,
            "checked_at":
                datetime.now(
                    IST
                ).isoformat(),
        }

        if existing_rows:
            response = requests.patch(
                base_url,
                headers={
                    **headers,
                    "Prefer":
                        "return=minimal",
                },
                params={
                    "challenge_id":
                        f"eq.{challenge_id}",
                    "register_number":
                        f"eq.{register_number}",
                },
                json=payload,
                timeout=30,
            )
        else:
            response = requests.post(
                base_url,
                headers={
                    **headers,
                    "Prefer":
                        "return=minimal",
                },
                json={
                    "challenge_id":
                        challenge_id,
                    "register_number":
                        register_number,
                    **payload,
                },
                timeout=30,
            )

        if not response.ok:
            print(
                "  Challenge save warning "
                f"[{register_number}]: "
                f"{response.status_code} "
                f"{response.text[:180]}"
            )
            return False

        return True

    except requests.RequestException as error:
        print(
            "  Challenge network warning "
            f"[{register_number}]: "
            f"{error}"
        )
        return False

    except Exception as error:
        print(
            "  Challenge unexpected warning "
            f"[{register_number}]: "
            f"{error}"
        )
        return False


# ============================================================
# PARALLEL STUDENT WORKER
# ============================================================

def previous_good_row(
    previous_history: pd.DataFrame,
    register_number: str,
) -> dict[str, Any] | None:
    """
    Latest usable history row for failure-safe display.
    """
    if previous_history is None or previous_history.empty:
        return None

    rows = previous_history[
        previous_history[
            "Register Number"
        ].astype(str) == str(register_number)
    ].copy()

    if rows.empty:
        return None

    if "Status" in rows.columns:
        success_rows = rows[
            rows["Status"].astype(str)
            .str.startswith("Success")
        ]

        if not success_rows.empty:
            rows = success_rows

    rows["_date"] = pd.to_datetime(
        rows["Date"],
        errors="coerce",
    )

    rows = rows.dropna(
        subset=["_date"]
    )

    if rows.empty:
        return None

    return (
        rows
        .sort_values("_date")
        .iloc[-1]
        .to_dict()
    )


def stale_profile_from_history(
    previous_history: pd.DataFrame,
    register_number: str,
    error_status: str,
) -> dict[str, Any]:
    """
    Never replace a previously good profile with all-zero values
    just because LeetCode temporarily failed.
    """
    old = previous_good_row(
        previous_history,
        register_number,
    )

    if old is None:
        return empty_profile(error_status)

    def number(name: str) -> int:
        return safe_int(old.get(name, 0))

    return {
        "total_solved":
            number("Problems Solved"),
        "easy":
            number("Easy"),
        "medium":
            number("Medium"),
        "hard":
            number("Hard"),
        "submissions":
            number("Total Submissions"),
        "solved_today":
            number("Solved Today"),
        "last_7_days":
            number("Last 7 Days"),
        "last_14_days":
            number("Last 14 Days"),
        "last_30_days":
            number("Last 30 Days"),
        "last_7_days_submissions":
            number("Last 7 Days Submissions"),
        "last_problem":
            clean(old.get("Last Problem", "")),
        "last_solved":
            clean(old.get("Last Solved", "")),
        "status":
            f"{error_status} | Previous data kept",
        "recent_submissions": [],
        "window_coverage": {
            "today": False,
            "7d": False,
            "14d": False,
            "30d": False,
        },
        "accepted_submission_total": 0,
        "recent_accepted_returned": 0,
    }


def validate_student_output(
    row: dict[str, Any],
) -> None:
    """
    Reject impossible student metrics before they reach LiveData/History/
    Supabase/frontend.
    """
    total = max(
        0,
        safe_int(row.get("Problems Solved")),
    )

    today = max(
        0,
        safe_int(row.get("Solved Today")),
    )

    seven = max(
        0,
        safe_int(row.get("Last 7 Days")),
    )

    fourteen = max(
        0,
        safe_int(row.get("Last 14 Days")),
    )

    thirty = max(
        0,
        safe_int(row.get("Last 30 Days")),
    )

    easy = max(0, safe_int(row.get("Easy")))
    medium = max(0, safe_int(row.get("Medium")))
    hard = max(0, safe_int(row.get("Hard")))

    if not (
        today
        <= seven
        <= fourteen
        <= thirty
        <= total
    ):
        raise RuntimeError(
            "Impossible rolling metrics: "
            f"T={today}, 7={seven}, 14={fourteen}, "
            f"30={thirty}, total={total}"
        )

    if total > 0 and (easy + medium + hard) != total:
        raise RuntimeError(
            "Difficulty total mismatch: "
            f"E={easy}, M={medium}, H={hard}, total={total}"
        )


def process_student(
    position: int,
    total_students: int,
    student: pd.Series,
    updated_at: str,
    previous_history: pd.DataFrame,
    previous_activity: pd.DataFrame,
    recent_challenges: list[dict[str, Any]],
) -> dict[str, Any]:
    register_number = clean(
        student["Register Number"]
    )

    student_name = clean(
        student["Student Name"]
    )

    username = clean(
        student["LeetCode Username"]
    )

    section = clean(
        student["Section"]
    )

    print(
        f"[START {position}/{total_students}] "
        f"{section} | "
        f"{student_name} ({username})"
    )

    profile = fetch_leetcode(username)

    fetch_success = (
        profile.get("status") == "Success"
    )

    if not fetch_success:
        fetch_error = clean(
            profile.get(
                "status",
                "LeetCode fetch failed",
            )
        )

        profile = stale_profile_from_history(
            previous_history,
            register_number,
            fetch_error,
        )

    if fetch_success:
        for challenge in recent_challenges:
            challenge_done, challenge_done_at = (
                challenge_completion(
                    profile.get(
                        "recent_submissions",
                        [],
                    ),
                    challenge,
                )
            )

            challenge_saved = (
                save_challenge_result(
                    int(challenge["id"]),
                    register_number,
                    challenge_done,
                    challenge_done_at,
                )
            )

            if challenge_done:
                print(
                    f"[CHALLENGE {position}/{total_students}] "
                    f"{student_name} | "
                    f"{challenge.get('problem_title', '')} | "
                    f"{'saved ✅' if challenge_saved else 'save warning ⚠️'}"
                )

    if fetch_success:
        rolling = calculate_rolling_metrics(
            previous_history,
            register_number,
            username,
            profile,
        )
    else:
        rolling = {
            "today": safe_int(profile.get("solved_today")),
            "7d": safe_int(profile.get("last_7_days")),
            "14d": safe_int(profile.get("last_14_days")),
            "30d": safe_int(profile.get("last_30_days")),
            "today_source": "STALE",
            "7d_source": "STALE",
            "14d_source": "STALE",
            "30d_source": "STALE",
        }

    completed_solved = rolling["today"]
    completed_7_days = rolling["7d"]
    completed_14_days = rolling["14d"]
    completed_30_days = rolling["30d"]

    # Calendar permission is independent from solved tracking.
    # Keep the previous 7D submission value when only the calendar is hidden.
    weekly_submissions = safe_int(
        profile.get(
            "last_7_days_submissions",
            0,
        )
    )

    calendar_available = profile.get(
        "calendar_available",
        True,
    )

    if fetch_success and not calendar_available:
        old = previous_good_row(
            previous_history,
            register_number,
        )

        if old is not None:
            weekly_submissions = safe_int(
                old.get(
                    "Last 7 Days Submissions",
                    weekly_submissions,
                )
            )

    completed_date = (
        ist_today().isoformat()
        if fetch_success
        else ""
    )

    source_note = (
        f"T:{rolling['today_source']} "
        f"7:{rolling['7d_source']} "
        f"14:{rolling['14d_source']} "
        f"30:{rolling['30d_source']}"
    )

    row = {
        "Section": section,
        "Register Number": register_number,
        "Student Name": student_name,
        "LeetCode Username": username,
        "LeetCode Link":
            f"https://leetcode.com/u/{username}/",
        "Problems Solved":
            profile["total_solved"],
        "Solved Today":
            completed_solved,
        "Last 7 Days":
            completed_7_days,
        "Last 14 Days":
            completed_14_days,
        "Last 30 Days":
            completed_30_days,
        "Last 7 Days Submissions":
            weekly_submissions,
        "Total Submissions":
            profile["submissions"],
        "Easy":
            profile["easy"],
        "Medium":
            profile["medium"],
        "Hard":
            profile["hard"],
        "Last Problem":
            profile["last_problem"],
        "Last Solved":
            profile["last_solved"],
        "Status":
            profile["status"],
        "Updated At":
            updated_at,
        "7D Source":
            rolling["7d_source"],
        "14D Source":
            rolling["14d_source"],
        "30D Source":
            rolling["30d_source"],
        "_Completed Date":
            completed_date,
        "_Completed Solved":
            completed_solved,
        "_Completed Source":
            rolling["today_source"],
        "_Fresh Success":
            fetch_success,
        "_Coverage":
            source_note,
    }

    validate_student_output(row)

    calendar_note = (
        ""
        if calendar_available
        else " | calendar unavailable; 7D submissions preserved"
    )

    print(
        f"[DONE  {position}/{total_students}] "
        f"{section} | "
        f"{student_name} | "
        f"30d={completed_30_days} | "
        f"14d={completed_14_days} | "
        f"7d={completed_7_days} | "
        f"sub7d={weekly_submissions} | "
        f"today={completed_solved} | "
        f"total={profile['total_solved']} | "
        f"{profile['status']} | "
        f"sources={source_note}"
        f"{calendar_note}"
    )

    return row


# ============================================================
# RANKING
# ============================================================

def add_ranks(
    live_data: pd.DataFrame,
) -> pd.DataFrame:
    if live_data.empty:
        return live_data

    ranking_columns = [
        "Last 30 Days",
        "Last 14 Days",
        "Last 7 Days",
        "Solved Today",
        "Problems Solved",
        "Student Name",
    ]

    ranking_ascending = [
        False,
        False,
        False,
        False,
        False,
        True,
    ]

    for metric_column in ranking_columns[:-1]:
        live_data[metric_column] = pd.to_numeric(
            live_data[metric_column],
            errors="coerce",
        ).fillna(0).astype(int)

    # --------------------------------------------------------
    # OVERALL RANK
    # --------------------------------------------------------

    overall_sorted = (
        live_data
        .sort_values(
            by=ranking_columns,
            ascending=ranking_ascending,
        )
        .reset_index(drop=True)
    )

    overall_sorted[
        "Overall Rank"
    ] = range(
        1,
        len(overall_sorted) + 1,
    )

    overall_rank_map = dict(
        zip(
            overall_sorted[
                "Register Number"
            ].astype(str),
            overall_sorted[
                "Overall Rank"
            ],
        )
    )

    live_data[
        "Overall Rank"
    ] = (
        live_data[
            "Register Number"
        ]
        .astype(str)
        .map(overall_rank_map)
    )

    # --------------------------------------------------------
    # SECTION RANK
    # --------------------------------------------------------

    section_rank_map: dict[
        tuple[str, str],
        int,
    ] = {}

    for (
        section,
        section_frame,
    ) in live_data.groupby(
        "Section",
        sort=False,
    ):
        section_sorted = (
            section_frame
            .sort_values(
                by=ranking_columns,
                ascending=ranking_ascending,
            )
            .reset_index(drop=True)
        )

        for index, row in (
            section_sorted
            .iterrows()
        ):
            key = (
                str(section),
                str(
                    row[
                        "Register Number"
                    ]
                ),
            )

            section_rank_map[
                key
            ] = index + 1

    live_data[
        "Section Rank"
    ] = live_data.apply(
        lambda row:
            section_rank_map.get(
                (
                    str(
                        row["Section"]
                    ),
                    str(
                        row[
                            "Register Number"
                        ]
                    ),
                ),
                "",
            ),
        axis=1,
    )

    # Store CSV in overall-rank order.
    live_data = (
        live_data
        .sort_values(
            by=[
                "Overall Rank"
            ],
            ascending=[True],
        )
        .reset_index(drop=True)
    )

    columns = [
        "Overall Rank",
        "Section Rank",
        "Section",
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "LeetCode Link",
        "Problems Solved",
        "Solved Today",
        "Last 7 Days",
        "Last 14 Days",
        "Last 30 Days",
        "Last 7 Days Submissions",
        "Total Submissions",
        "Easy",
        "Medium",
        "Hard",
        "Last Problem",
        "Last Solved",
        "Status",
        "Updated At",
    ]

    return live_data[columns]


# ============================================================
# HISTORY
# ============================================================

HISTORY_COLUMNS = [
    "Date",
    "Overall Rank",
    "Section Rank",
    "Section",
    "Register Number",
    "Student Name",
    "LeetCode Username",
    "Problems Solved",
    "Solved Today",
    "Last 7 Days",
    "Last 14 Days",
    "Last 30 Days",
    "Last 7 Days Submissions",
    "Total Submissions",
    "Easy",
    "Medium",
    "Hard",
    "Last Problem",
    "Last Solved",
    "Status",
    "Updated At",
]


def load_history() -> pd.DataFrame:
    if not HISTORY_CSV.exists():
        return pd.DataFrame(
            columns=HISTORY_COLUMNS
        )

    try:
        history = pd.read_csv(
            HISTORY_CSV,
            dtype=str,
            keep_default_na=False,
        )
    except pd.errors.EmptyDataError:
        return pd.DataFrame(
            columns=HISTORY_COLUMNS
        )

    for column in HISTORY_COLUMNS:
        if column not in history.columns:
            history[column] = ""

    return history[
        HISTORY_COLUMNS
    ].copy()


def update_history(
    previous_history: pd.DataFrame,
    current_rows: list[
        dict[str, Any]
    ],
) -> pd.DataFrame:
    today_text = ist_today().isoformat()

    history = (
        previous_history.copy()
    )

    if not history.empty:
        current_registers = {
            str(
                row[
                    "Register Number"
                ]
            )
            for row in current_rows
        }

        history = history[
            ~(
                (
                    history["Date"]
                    == today_text
                )
                & (
                    history[
                        "Register Number"
                    ]
                    .astype(str)
                    .isin(
                        current_registers
                    )
                )
            )
        ].copy()

    new_history_rows = []

    for row in current_rows:
        new_history_rows.append(
            {
                "Date":
                    today_text,
                "Overall Rank":
                    row.get(
                        "Overall Rank",
                        "",
                    ),
                "Section Rank":
                    row.get(
                        "Section Rank",
                        "",
                    ),
                "Section":
                    row.get(
                        "Section",
                        "",
                    ),
                "Register Number":
                    row[
                        "Register Number"
                    ],
                "Student Name":
                    row[
                        "Student Name"
                    ],
                "LeetCode Username":
                    row[
                        "LeetCode Username"
                    ],
                "Problems Solved":
                    row[
                        "Problems Solved"
                    ],
                "Solved Today":
                    row[
                        "Solved Today"
                    ],
                "Last 7 Days":
                    row[
                        "Last 7 Days"
                    ],
                "Last 14 Days":
                    row.get(
                        "Last 14 Days",
                        0,
                    ),
                "Last 30 Days":
                    row[
                        "Last 30 Days"
                    ],
                "Last 7 Days Submissions":
                    row.get(
                        "Last 7 Days Submissions",
                        0,
                    ),
                "Total Submissions":
                    row[
                        "Total Submissions"
                    ],
                "Easy":
                    row["Easy"],
                "Medium":
                    row["Medium"],
                "Hard":
                    row["Hard"],
                "Last Problem":
                    row[
                        "Last Problem"
                    ],
                "Last Solved":
                    row[
                        "Last Solved"
                    ],
                "Status":
                    row["Status"],
                "Updated At":
                    row[
                        "Updated At"
                    ],
            }
        )

    combined = pd.concat(
        [
            history,
            pd.DataFrame(
                new_history_rows,
                columns=HISTORY_COLUMNS,
            ),
        ],
        ignore_index=True,
    )

    if not combined.empty:
        combined = (
            combined
            .sort_values(
                by=[
                    "Date",
                    "Section",
                    "Student Name",
                ],
                ascending=[
                    True,
                    True,
                    True,
                ],
            )
            .reset_index(
                drop=True
            )
        )

    return combined


def build_daily_activity(
    history: pd.DataFrame,
) -> pd.DataFrame:
    columns = [
        "Date",
        "Section",
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "Problems Solved",
        "Solved That Day",
    ]

    if history.empty:
        return pd.DataFrame(
            columns=columns
        )

    activity = history[
        [
            "Date",
            "Section",
            "Register Number",
            "Student Name",
            "LeetCode Username",
            "Problems Solved",
            "Solved Today",
        ]
    ].copy()

    activity = activity.rename(
        columns={
            "Solved Today":
                "Solved That Day"
        }
    )

    return activity[
        columns
    ]



# ============================================================
# AI PERFORMANCE ANALYST — SUPABASE DATA MART SYNC
# ============================================================
def _postgrest_upsert(
    table: str,
    rows: list[dict[str, Any]],
    on_conflict: str,
) -> None:
    if not rows:
        return

    url = (
        f"{SUPABASE_URL}/rest/v1/{table}"
        f"?on_conflict={on_conflict}"
    )

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }

    # Chunk payloads to stay comfortably below request-size limits.
    chunk_size = 150

    for start in range(0, len(rows), chunk_size):
        response = requests.post(
            url,
            headers=headers,
            json=rows[start:start + chunk_size],
            timeout=45,
        )

        response.raise_for_status()


def sync_ai_performance_tables(live_data: pd.DataFrame) -> None:
    """Publish deterministic tracker metrics to Supabase for the AI analyst."""

    if (
        not SUPABASE_URL
        or not SUPABASE_SERVICE_ROLE_KEY
        or live_data.empty
    ):
        return

    now = datetime.now(IST)
    snapshot_date = now.date().isoformat()
    updated_at = now.isoformat()

    current_rows: list[dict[str, Any]] = []
    history_rows: list[dict[str, Any]] = []

    def number(row: pd.Series, column: str) -> int:
        try:
            value = row.get(column, 0)
            if pd.isna(value):
                return 0
            return int(float(value))
        except (TypeError, ValueError):
            return 0

    for _, row in live_data.iterrows():
        register_number = clean(row.get("Register Number", ""))

        if not register_number:
            continue

        current_rows.append({
            "register_number": register_number,
            "student_name": clean(row.get("Student Name", "")),
            "section": clean(row.get("Section", "")),
            "leetcode_username": clean(row.get("LeetCode Username", "")),
            "total_solved": number(row, "Problems Solved"),
            "solved_today": number(row, "Solved Today"),
            "last_7_days": number(row, "Last 7 Days"),
            "last_14_days": number(row, "Last 14 Days"),
            "last_30_days": number(row, "Last 30 Days"),
            "last_7_days_submissions":
                number(row, "Last 7 Days Submissions"),
            "total_submissions": number(row, "Total Submissions"),
            "easy": number(row, "Easy"),
            "medium": number(row, "Medium"),
            "hard": number(row, "Hard"),
            "last_problem": clean(row.get("Last Problem", "")),
            "last_solved": clean(row.get("Last Solved", "")),
            "status": clean(row.get("Status", "Pending")) or "Pending",
            "overall_rank": number(row, "Overall Rank") or None,
            "section_rank": number(row, "Section Rank") or None,
            "updated_at": updated_at,
        })

        if clean(row.get("Status", "")) == "Success":
            history_rows.append({
                    "register_number": register_number,
                    "snapshot_date": snapshot_date,
                    "total_solved": number(row, "Problems Solved"),
                "solved_today": number(row, "Solved Today"),
                "last_7_days": number(row, "Last 7 Days"),
                "last_14_days": number(row, "Last 14 Days"),
                "last_30_days": number(row, "Last 30 Days"),
                "last_7_days_submissions":
                    number(row, "Last 7 Days Submissions"),
                "easy": number(row, "Easy"),
                "medium": number(row, "Medium"),
                "hard": number(row, "Hard"),
                "updated_at": updated_at,
            })

    try:
        _postgrest_upsert(
            "student_performance_current",
            current_rows,
            "register_number",
        )

        _postgrest_upsert(
            "student_performance_history",
            history_rows,
            "register_number,snapshot_date",
        )

        print(
            f"AI analytics sync: {len(current_rows)} current profiles "
            f"+ {len(history_rows)} daily snapshots."
        )

    except Exception as error:
        # The normal CSV tracker must continue even when the optional AI mart is unavailable.
        print(f"[AI ANALYTICS SYNC WARNING] {error}")

# ============================================================
# MAIN UPDATE
# ============================================================

def run_one_update() -> None:
    students = read_students()

    if not students.empty:
        duplicated_registers = students[
            students["Register Number"]
            .astype(str)
            .duplicated(keep=False)
        ]

        if not duplicated_registers.empty:
            print(
                "[WARNING] Duplicate register numbers found; "
                "keeping the last row for each register number."
            )

            students = (
                students
                .drop_duplicates(
                    subset=["Register Number"],
                    keep="last",
                )
                .reset_index(drop=True)
            )

    updated_at = datetime.now(IST).strftime(
        "%Y-%m-%d %H:%M:%S IST"
    )

    live_rows: list[
        dict[str, Any]
    ] = []

    total_students = len(
        students
    )

    worker_count = (
        min(
            MAX_WORKERS,
            total_students,
        )
        if total_students
        else 1
    )

    print("=" * 64)
    print(
        f"LeetCode cloud update: "
        f"{updated_at}"
    )
    print(
        f"Students: "
        f"{total_students}"
    )
    # Load shared data BEFORE printing/using it.
    previous_history = load_history()
    previous_activity = load_daily_activity_file()

    # Load challenge definitions once.
    # The same read-only list is shared safely across all workers.
    recent_challenges = load_recent_challenges()

    print(
        f"Parallel workers: "
        f"{worker_count}"
    )
    print(
        f"Daily challenges loaded: "
        f"{len(recent_challenges)}"
    )
    print("=" * 64)

    futures = {}

    with ThreadPoolExecutor(
        max_workers=worker_count
    ) as executor:
        for (
            position,
            (_, student),
        ) in enumerate(
            students.iterrows(),
            start=1,
        ):
            future = executor.submit(
                process_student,
                position,
                total_students,
                student,
                updated_at,
                previous_history,
                previous_activity,
                recent_challenges,
            )

            futures[
                future
            ] = {
                "position":
                    position,
                "section":
                    clean(
                        student[
                            "Section"
                        ]
                    ),
                "student_name":
                    clean(
                        student[
                            "Student Name"
                        ]
                    ),
                "register_number":
                    clean(
                        student[
                            "Register Number"
                        ]
                    ),
                "username":
                    clean(
                        student[
                            "LeetCode Username"
                        ]
                    ),
            }

        for future in as_completed(
            futures
        ):
            info = futures[
                future
            ]

            try:
                live_rows.append(
                    future.result()
                )

            except Exception as error:
                print(
                    f"[FAILED "
                    f"{info['position']}/"
                    f"{total_students}] "
                    f"{info['section']} | "
                    f"{info['student_name']} "
                    f"({info['username']}): "
                    f"{error}"
                )

                stale = stale_profile_from_history(
                    previous_history,
                    info["register_number"],
                    f"Worker error: {error}",
                )

                live_rows.append(
                    {
                        "Section":
                            info["section"],
                        "Register Number":
                            info["register_number"],
                        "Student Name":
                            info["student_name"],
                        "LeetCode Username":
                            info["username"],
                        "LeetCode Link":
                            (
                                "https://leetcode.com/u/"
                                f"{info['username']}/"
                            ),
                        "Problems Solved":
                            stale["total_solved"],
                        "Solved Today":
                            stale["solved_today"],
                        "Last 7 Days":
                            stale["last_7_days"],
                        "Last 14 Days":
                            stale["last_14_days"],
                        "Last 30 Days":
                            stale["last_30_days"],
                        "Last 7 Days Submissions":
                            stale["last_7_days_submissions"],
                        "Total Submissions":
                            stale["submissions"],
                        "Easy":
                            stale["easy"],
                        "Medium":
                            stale["medium"],
                        "Hard":
                            stale["hard"],
                        "Last Problem":
                            stale["last_problem"],
                        "Last Solved":
                            stale["last_solved"],
                        "Status":
                            stale["status"],
                        "Updated At":
                            updated_at,
                        "_Completed Date":
                            "",
                        "_Completed Solved":
                            0,
                        "_Fresh Success":
                            False,
                        "_Coverage":
                            "worker failure",
                    }
                )

    # Keep daily activity/history only from fresh successful fetches.
    completed_activity_rows = [
        row
        for row in live_rows
        if bool(row.get("_Fresh Success"))
    ]

    fresh_history_rows = [
        {
            key: value
            for key, value in row.items()
            if not key.startswith("_")
        }
        for row in live_rows
        if bool(row.get("_Fresh Success"))
    ]

    live_data = pd.DataFrame(
        live_rows
    )

    # Internal helper columns must not appear in LiveData.csv.
    live_data = live_data.drop(
        columns=[
            "_Completed Date",
            "_Completed Solved",
            "_Fresh Success",
            "_Coverage",
            "_Completed Source",
        ],
        errors="ignore",
    )

    live_data = add_ranks(
        live_data
    )

    # Never replace a good daily history snapshot with zeros/stale values
    # from a failed LeetCode request.
    history = update_history(
        previous_history,
        fresh_history_rows,
    )

    daily_activity = update_completed_daily_activity(
        previous_activity,
        completed_activity_rows,
    )

    atomic_csv_write(
        live_data,
        LIVE_CSV,
    )

    atomic_csv_write(
        history,
        HISTORY_CSV,
    )

    atomic_csv_write(
        daily_activity,
        DAILY_ACTIVITY_CSV,
    )

    sync_ai_performance_tables(
        live_data,
    )

    print("=" * 64)
    print(
        "CSV files updated "
        "successfully."
    )
    print(
        f"LiveData.csv: "
        f"{LIVE_CSV}"
    )
    print(
        f"History.csv: "
        f"{HISTORY_CSV}"
    )
    print(
        f"DailyActivity.csv: "
        f"{DAILY_ACTIVITY_CSV}"
    )
    print("=" * 64)


if __name__ == "__main__":
    run_one_update()
